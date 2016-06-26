"""
file-to-db
readfile.py

read in data file and generate data object for further db queries

@By Seth (Xiaohui) Wang
@email: sethwang199418@gmail.com
"""
import csv
import sys
from query import *
from db import *
from openpyxl import load_workbook
import urllib2
import urllib
import zipfile

formats = ['csv', 'txt', 'json']
supported_types = ['text', 'integer']

def check_format(filename):
    tokens = filename.split('.')
    return tokens[-1]

######################################################
####           entry to parse file                ###
######################################################
def parse_file(filePath, table_name, primary_key = None):
    f_format = check_format(filePath)
    if f_format in formats:
        if f_format == 'csv':
            parse_csv(filePath, table_name, primary_key)

######################################################
####  Parse .csv file and return db query object   ###
######################################################
def parse_csv(file_path, table_name, primary_key=None):
    # for remote files
    f = urllib2.urlopen(file_path)
    # for local files
    #f = open(filename)
    data = []
    types = {}
    col_names = []
    try:
        i = 0
        col_names = []
        reader = csv.reader(f.read().splitlines())
        #reader = csv.reader(f.read().splitlines(), dialect=csv.excel_tab)
        for row in reader:
            if i == 0:
                col_names = row
                i+=1
                print(col_names)
            elif i == 1:
                try:
                    for j in range(len(row)):
                        if row[j].isdigit():
                            types[col_names[j]] = 'integer'
                        else:
                            types[col_names[j]] = 'text'
                    data.append(row)
                    assert(len(row) == len(col_names))
                except:
                    print "Data does not have equal number of column as column names"
                i += 1
            else:
                data.append(row)
    except Exception as e:
        print e
    #print(col_names)
    f.close()
    q = query(table_name, primary_key, col_names, data, types)
    return q

def parse_xl(file_path, table_name, primary_key=None):
    #f = urllib2.urlopen(file_name)
    file_name = file_path.split('/')[-1]
    #file_suffix = file_path.split('/')[-1]
    #if file_suffix.split(".")[1] == "xls":
    #    file_name += "x"
    f = urllib.urlretrieve(file_path, file_name)

    wb = load_workbook(file_name, read_only=True)
    # local files
    #wb = load_workbook(file_name, read_only= True)
    sheets = wb.get_sheet_names()
    print(sheets)
    queries = []
    ws = None
    types = {}
    for sheet in sheets:
        col_names = []
        sheet_data = []
        ws = wb[sheet]
        i = 0
        for row in ws.rows:
            if i == 0:
                col_names = [cell.value for cell in row]
                i += 1
            elif i == 1:
                for j in range(len(row)):

                    if cell.value.isdigit():
                        types[col_names[j]] = 'integer'
                    else:
                        types[col_names[j]] = 'text'
                i += 1
            else:
                row_data = [cell.value for cell in row]
                sheet_data.append(row_data)
        q = query(sheet, None, col_names, sheet_data, types)
        queries.append(q)
    f.close()
    return queries
